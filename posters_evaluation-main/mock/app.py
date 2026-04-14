"""
Test/Mock run script that simulates OpenAI API calls using mock results
Run this instead of run.py to test the GUI without OpenAI API key

Usage:
    python test_run.py
"""
import sys
import tkinter as tk
from pathlib import Path
from typing import Dict

# Add mock directory to path
sys.path.insert(0, str(Path(__file__).parent / "mock"))

from mock.generate_mock_results import populate_folder
from gui.app import PosterEvaluationGUI


# Mock data storage
MOCK_JOBS = {}
DOWNLOADS_DIR = Path(__file__).parent / "downloads"


def mock_upload_batch(self, folder_path: str, approach: str = "direct"):
    """Mock file upload that generates results"""
    print(f"[MOCK] Uploading batch with approach: {approach}")
    
    folder = Path(folder_path)
    image_extensions = {'.jpg', '.jpeg', '.png'}
    image_files = [f for f in folder.iterdir() if f.suffix.lower() in image_extensions]
    
    if not image_files:
        return None
    
    # Generate mock job ID
    import uuid
    job_id = f"mock-job-{str(uuid.uuid4())[:12]}"
    
    # Generate results
    poster_files = [f.name for f in image_files]
    job_folder = DOWNLOADS_DIR / job_id
    job_folder.mkdir(parents=True, exist_ok=True)
    
    populate_folder(job_folder, approach, poster_files)
    
    # Store job info
    MOCK_JOBS[job_id] = {
        'approach': approach,
        'status': 'completed',
        'folder': str(job_folder),
        'poster_files': poster_files
    }
    
    print(f"[MOCK] Generated results for job {job_id}")
    return job_id


def mock_poll_job_status(self, job_id: str):
    """Mock job status polling"""
    if job_id in MOCK_JOBS:
        return {
            'status': 'completed',
            'progress': 100,
            'job_id': job_id
        }
    return {
        'status': 'error',
        'message': f'Job {job_id} not found'
    }


def mock_get_results(self, job_id: str):
    """Mock results retrieval from CSV"""
    if job_id not in MOCK_JOBS:
        return None
    
    job_info = MOCK_JOBS[job_id]
    job_folder = Path(job_info['folder'])
    results_csv = job_folder / "results_master.csv"
    
    if not results_csv.exists():
        return None
    
    # Parse CSV results
    import csv
    results = []
    with open(results_csv, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Convert grades to float
            result = {k: v for k, v in row.items()}
            result['final_grade'] = float(row['final_grade'])
            results.append(result)
    
    return {
        'job_id': job_id,
        'status': 'completed',
        'results': results
    }


def mock_download_excel(self, job_id: str, save_path: str):
    """Mock Excel download with results table"""
    if job_id not in MOCK_JOBS:
        return False
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import csv
        
        job_info = MOCK_JOBS[job_id]
        job_folder = Path(job_info['folder'])
        results_csv = job_folder / "results_master.csv"
        
        if not results_csv.exists():
            return False
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Evaluation Results"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Add title
        ws['A1'] = "Academic Poster Evaluation Results"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        # Add headers
        headers = ['Project Number', 'Publisher Names', 'Direct Grade', 'Reasoning Grade', 'Deep Analysis Grade', 'Strict Grade']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Read CSV and add data
        row_idx = 4
        with open(results_csv, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                ws.cell(row=row_idx, column=1).value = row.get('project_number', '')
                ws.cell(row=row_idx, column=2).value = row.get('presenter_names', '')
                ws.cell(row=row_idx, column=3).value = float(row.get('final_grade', 0))
                ws.cell(row=row_idx, column=4).value = float(row.get('final_grade', 0))
                ws.cell(row=row_idx, column=5).value = float(row.get('final_grade', 0))
                ws.cell(row=row_idx, column=6).value = float(row.get('final_grade', 0))
                
                # Apply borders and center alignment to data cells
                for col in range(1, 7):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = border
                    if col >= 3:  # Center align grade columns
                        cell.alignment = center_alignment
                
                row_idx += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15
        
        # Save file
        try:
            wb.save(save_path)
            print(f"[MOCK] Excel file saved to {save_path}")
            return True
        except PermissionError:
            print(f"[MOCK] Permission denied: Could not save to {save_path}. Is the file open in another program?")
            # We can't use messagebox here directly because we are in a thread usually, 
            # but in this mock it might be called differently. 
            # Actually, the original code returns False which the GUI then handles.
            return False
    except ImportError:
        # Fallback: create CSV-like output
        try:
            job_info = MOCK_JOBS[job_id]
            job_folder = Path(job_info['folder'])
            results_csv = job_folder / "results_master.csv"
            
            # Convert CSV to simple text table in Excel-like format
            import shutil
            save_path_csv = save_path.replace('.xlsx', '.csv')
            shutil.copy(results_csv, save_path_csv)
            print(f"[MOCK] Results saved as CSV: {save_path_csv}")
            return True
        except Exception as e:
            print(f"[MOCK] Error saving file: {e}")
            return False


def mock_download_comparison_excel(self, job_ids: Dict[str, str], save_path: str) -> bool:
    """Mock comparison Excel download"""
    print(f"[MOCK] Generating comparison Excel for {len(job_ids)} jobs")
    
    # In mock mode, we can just use the download_excel mock for any of the job IDs
    # since mock results are consistent across approaches
    if not job_ids:
        return False
        
    first_job_id = list(job_ids.values())[0]
    return mock_download_excel(self, first_job_id, save_path)


def mock_start_server(api_key: str = None):
    """Mock server startup"""
    print("[MOCK] Server is running in mock mode")
    return True


def mock_is_running():
    """Mock server health check"""
    return True


def setup_mocks():
    """Setup all mock functions"""
    print("\n" + "="*70)
    print(" "*15 + "MOCK MODE - SIMULATING OPENAI API CALLS")
    print("="*70)
    print("✓ All HTTP requests are MOCKED")
    print("✓ Results from generate_mock_results.py are used")
    print("✓ No OpenAI API key needed")
    print("✓ No actual FastAPI server needed")
    print("="*70 + "\n")
    
    # Patch the backend classes
    from gui import backend
    
    backend.EvaluationClient.upload_batch = mock_upload_batch
    backend.EvaluationClient.poll_job_status = mock_poll_job_status
    backend.EvaluationClient.get_results = mock_get_results
    backend.EvaluationClient.download_excel = mock_download_excel
    backend.EvaluationClient.download_comparison_excel = mock_download_comparison_excel
    
    backend.ServerManager.start_server = mock_start_server
    backend.ServerManager.is_running = lambda self: mock_is_running()
    backend.ServerManager.stop_server = lambda self: None  # No-op


if __name__ == "__main__":
    # Setup mock mode
    setup_mocks()
    
    # Create downloads directory if needed
    DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
    
    # Run GUI with mocks
    try:
        root = tk.Tk()
        app = PosterEvaluationGUI(root)
        root.mainloop()
    except KeyboardInterrupt:
        print("\nShutting down...")
        sys.exit(0)

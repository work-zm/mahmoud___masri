import json
import asyncio
import aiofiles
import pandas as pd
from pathlib import Path
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ..models.poster_data import PosterEvaluation, ProcessingLog

class AsyncOutputGenerator:
    """Generate all required download files asynchronously"""
    
    
    def __init__(self, download_dir: Path):
        self.download_dir = download_dir
        self.download_dir.mkdir(parents=True, exist_ok=True)
    
    async def generate_master_results(self, evaluations: List[PosterEvaluation]) -> Path:
        """Generate master CSV results file"""
        filename = f"results_master.csv"
        filepath = self.download_dir / filename
        
        # Prepare data for CSV
        data = []
        for eval in evaluations:
            data.append({
                "Poster File": eval.poster_file,
                "Final Grade": eval.final_grade,
                "Project Number": eval.project_number if eval.project_number else "",
                "Project Summary": eval.poster_summary,
                "Evaluation Summary": eval.evaluation_summary
            })
        
        # Create DataFrame and save
        df = pd.DataFrame(data)
        
        # Use asyncio to write file
        csv_content = df.to_csv(index=False)
        async with aiofiles.open(filepath, 'w', encoding='utf-8', newline='') as f:
            await f.write(csv_content.strip())
        
        print(f"Master results saved to: {filepath}")
        return filepath
    
    async def generate_excel_results(self, evaluations: List[PosterEvaluation]) -> Path:
        """Generate Excel results file with formatted table"""
        filename = "results_comparison.xlsx"
        filepath = self.download_dir / filename
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Evaluation Results"
        
        # Define headers
        headers = ["Project Number", "Publisher Names", "Grade"]
        
        # Style definitions
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        grade_alignment = Alignment(horizontal="center", vertical="center")
        
        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row_idx, eval in enumerate(evaluations, start=2):
            # Project Number
            cell = ws.cell(row=row_idx, column=1, value=eval.project_number or "N/A")
            cell.alignment = cell_alignment
            cell.border = border
            
            # Publisher Names
            cell = ws.cell(row=row_idx, column=2, value=eval.presenter_names or "N/A")
            cell.alignment = cell_alignment
            cell.border = border
            
            # Grade
            cell = ws.cell(row=row_idx, column=3, value=eval.final_grade)
            cell.alignment = grade_alignment
            cell.border = border
            
            # Color code grades
            if eval.final_grade >= 80:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif eval.final_grade >= 60:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        
        # Save workbook
        wb.save(filepath)
        
        print(f"Excel results saved to: {filepath}")
        return filepath
    
    async def generate_comparison_excel(self, combined_results: List[dict]) -> Path:
        """Generate Excel results file comparing all 4 approaches"""
        filename = "results_comparison_all.xlsx"
        filepath = self.download_dir / filename
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparison Results"
        
        # Define headers (6 columns as in GUI)
        headers = ["Project Number", "Publisher Names", "Direct", "Reasoning", "Deep Analysis", "Strict"]
        
        # Style definitions
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        grade_alignment = Alignment(horizontal="center", vertical="center")
        
        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row_idx, result in enumerate(combined_results, start=2):
            # Project Number
            cell = ws.cell(row=row_idx, column=1, value=result.get("project_number", "N/A"))
            cell.alignment = cell_alignment
            cell.border = border
            
            # Publisher Names
            cell = ws.cell(row=row_idx, column=2, value=result.get("presenter_names", "N/A"))
            cell.alignment = cell_alignment
            cell.border = border
            
            # Grades (columns 3-6)
            grade_keys = ["direct_grade", "reasoning_grade", "deep_analysis_grade", "strict_grade"]
            for col_offset, key in enumerate(grade_keys):
                grade = result.get(key, 0)
                cell = ws.cell(row=row_idx, column=3 + col_offset, value=grade)
                cell.alignment = grade_alignment
                cell.border = border
                
                # Color code grades
                if grade >= 80:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif grade >= 60:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 10
        
        # Save workbook
        wb.save(filepath)
        
        print(f"Comparison Excel saved to: {filepath}")
        return filepath
    
    async def generate_individual_breakdowns(self, evaluations: List[PosterEvaluation]) -> List[Path]:
        """Generate individual JSON breakdown files"""
        breakdown_files = []
        
        async def write_breakdown_file(eval: PosterEvaluation):
            # Determine filename
            if eval.project_number and eval.presenter_names:  # Project number and presenter found
                filename = f"{eval.project_number}_{eval.presenter_names}.json"
            elif eval.project_number:
                filename = f"{eval.project_number}.json"
            else:
                # Fallback naming
                stem = Path(eval.poster_file).stem
                filename = f"{stem}_Unknown.json"
            
            # Clean filename (remove invalid characters)
            filename = "".join(c for c in filename if c.isalnum() or c in "._- ")
            filepath = self.download_dir / filename
            
            # Create JSON data
            json_data = eval.to_dict()
            
            # Write JSON file asynchronously
            async with aiofiles.open(filepath, 'w', encoding='utf-8') as f:
                await f.write(json.dumps(json_data, indent=2, ensure_ascii=False))
            
            return filepath
        
        # Write all breakdown files concurrently
        tasks = [write_breakdown_file(eval) for eval in evaluations]
        breakdown_files = await asyncio.gather(*tasks)
        
        print(f"Generated {len(breakdown_files)} breakdown files")
        return breakdown_files
    
    async def generate_run_log(self, logs: List[ProcessingLog]) -> Path:
        """Generate JSONL run log file"""
        filepath = self.download_dir / "run_log.jsonl"
        
        async with aiofiles.open(filepath, 'w', encoding='utf-8') as f:
            for log in logs:
                await f.write(log.json() + '\n')
        
        print(f"Run log saved to: {filepath}")
        return filepath
    
    async def generate_all_outputs(self, evaluations: List[PosterEvaluation], 
                                 logs: List[ProcessingLog]) -> dict:
        """Generate all download files concurrently"""
        
        # Run all generation tasks concurrently
        master_task = self.generate_master_results(evaluations)
        excel_task = self.generate_excel_results(evaluations)
        breakdown_task = self.generate_individual_breakdowns(evaluations)
        log_task = self.generate_run_log(logs)
        
        master_file, excel_file, breakdown_files, log_file = await asyncio.gather(
            master_task, excel_task, breakdown_task, log_task
        )
        
        return {
            "master_file": master_file,
            "excel_file": excel_file,
            "breakdown_files": breakdown_files,
            "log_file": log_file
        }
